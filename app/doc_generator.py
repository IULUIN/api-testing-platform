"""
测试用例文档生成器
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os


class TestCaseDocGenerator:
    """测试用例文档生成器"""

    def __init__(self):
        self.case_templates = {
            'positive': {
                'type': '功能测试',
                'priority': 'P0',
                'steps_template': [
                    '1. 准备测试数据',
                    '2. 执行{action}',
                    '3. 验证结果'
                ],
                'expected_template': '操作成功，{result}'
            },
            'negative': {
                'type': '异常测试',
                'priority': 'P1',
                'steps_template': [
                    '1. 准备异常测试数据',
                    '2. 执行{action}',
                    '3. 验证错误提示'
                ],
                'expected_template': '系统提示错误信息，{result}'
            },
            'boundary': {
                'type': '边界测试',
                'priority': 'P1',
                'steps_template': [
                    '1. 准备边界值数据',
                    '2. 执行{action}',
                    '3. 验证边界处理'
                ],
                'expected_template': '边界值处理正确，{result}'
            }
        }

    def generate_test_cases(self, data):
        """
        生成测试用例数据

        参数:
            data: {
                'scenario': '测试场景',
                'project_name': '项目名称',
                'tester': '测试人员',
                'features': [
                    {
                        'name': '一级功能点',
                        'sub_features': ['二级功能点1', '二级功能点2'],
                        'case_count': 2
                    }
                ],
                'include_steps': True,
                'include_priority': True
            }

        返回:
            测试用例列表
        """
        cases = []
        case_num = 1

        for feature in data['features']:
            feature_name = feature['name']
            sub_features = feature.get('sub_features', [])
            case_count = feature.get('case_count', 2)

            if not sub_features:
                # 没有二级功能点，直接生成
                for i in range(case_count):
                    case = self._generate_single_case(
                        case_num, feature_name, None, i, case_count, data
                    )
                    cases.append(case)
                    case_num += 1
            else:
                # 有二级功能点
                for sub_feature in sub_features:
                    for i in range(case_count):
                        case = self._generate_single_case(
                            case_num, feature_name, sub_feature, i, case_count, data
                        )
                        cases.append(case)
                        case_num += 1

        return cases

    def _generate_single_case(self, case_num, feature_name, sub_feature, index, total, data):
        """生成单个测试用例"""
        # 确定用例类型
        if index == 0:
            template_type = 'positive'
        elif index == 1:
            template_type = 'negative'
        else:
            template_type = 'boundary'

        template = self.case_templates[template_type]

        # 生成用例标题
        if sub_feature:
            title = f"{feature_name} - {sub_feature} - {template['type']}"
            full_feature = f"{feature_name} > {sub_feature}"
            action = sub_feature
        else:
            title = f"{feature_name} - {template['type']}"
            full_feature = feature_name
            action = feature_name

        # 生成测试步骤
        steps = []
        if data.get('include_steps', True):
            for step in template['steps_template']:
                steps.append(step.format(action=action))

        # 生成预期结果
        if template_type == 'positive':
            expected = template['expected_template'].format(result='符合预期')
        elif template_type == 'negative':
            expected = template['expected_template'].format(result='不允许操作')
        else:
            expected = template['expected_template'].format(result='边界值正确处理')

        # 生成前置条件
        if template_type == 'positive':
            precondition = '系统正常运行，用户已登录'
        else:
            precondition = '系统正常运行'

        case = {
            'id': f"TC{str(case_num).zfill(3)}",
            'title': title,
            'feature': full_feature,
            'type': template['type'],
            'priority': template['priority'] if data.get('include_priority', True) else '',
            'precondition': precondition,
            'steps': '\n'.join(steps) if steps else '',
            'expected': expected,
            'actual': '',
            'status': '',
            'remark': ''
        }

        return case

    def export_to_excel(self, cases, data, file_path=None):
        """
        导出为Excel文件

        参数:
            cases: 测试用例列表
            data: 配置数据
            file_path: 保存路径

        返回:
            文件路径
        """
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 20

        # 标题行
        title_row = 1
        ws.merge_cells(f'A{title_row}:K{title_row}')
        title_cell = ws[f'A{title_row}']
        title_cell.value = f"{data.get('scenario', '测试场景')} - 测试用例"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.font = Font(size=16, bold=True, color='FFFFFF')
        ws.row_dimensions[title_row].height = 30

        # 信息行
        info_row = 2
        info_text = f"项目：{data.get('project_name', 'N/A')}    测试人员：{data.get('tester', 'N/A')}    生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells(f'A{info_row}:K{info_row}')
        info_cell = ws[f'A{info_row}']
        info_cell.value = info_text
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        info_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        # 表头
        header_row = 3
        headers = ['用例编号', '用例标题', '功能点', '用例类型', '优先级', '前置条件', '测试步骤', '预期结果', '实际结果', '状态', '备注']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 数据行
        for row_idx, case in enumerate(cases, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=case['id'])
            ws.cell(row=row_idx, column=2, value=case['title'])
            ws.cell(row=row_idx, column=3, value=case['feature'])
            ws.cell(row=row_idx, column=4, value=case['type'])
            ws.cell(row=row_idx, column=5, value=case['priority'])
            ws.cell(row=row_idx, column=6, value=case['precondition'])
            ws.cell(row=row_idx, column=7, value=case['steps'])
            ws.cell(row=row_idx, column=8, value=case['expected'])
            ws.cell(row=row_idx, column=9, value=case['actual'])
            ws.cell(row=row_idx, column=10, value=case['status'])
            ws.cell(row=row_idx, column=11, value=case['remark'])

            # 设置样式
            for col in range(1, 12):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 设置行高
            ws.row_dimensions[row_idx].height = 60

        # 确定保存路径
        if not file_path:
            # 默认保存到项目的outputs目录
            output_dir = os.path.join(os.getcwd(), 'outputs')
            os.makedirs(output_dir, exist_ok=True)
            filename = f"测试用例_{data.get('scenario', '场景')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(output_dir, filename)

        # 保存文件
        wb.save(file_path)
        return file_path

    def export_to_markdown(self, cases, data, file_path=None):
        """
        导出为Markdown文件

        参数:
            cases: 测试用例列表
            data: 配置数据
            file_path: 保存路径

        返回:
            文件路径
        """
        lines = []

        # 标题
        lines.append(f"# {data.get('scenario', '测试场景')} - 测试用例\n")
        lines.append(f"**项目：** {data.get('project_name', 'N/A')}\n")
        lines.append(f"**测试人员：** {data.get('tester', 'N/A')}\n")
        lines.append(f"**生成时间：** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append(f"**用例总数：** {len(cases)}\n")
        lines.append("\n---\n\n")

        # 用例列表
        for case in cases:
            lines.append(f"## {case['id']} - {case['title']}\n\n")
            lines.append(f"**功能点：** {case['feature']}\n\n")
            lines.append(f"**用例类型：** {case['type']}\n\n")
            if case['priority']:
                lines.append(f"**优先级：** {case['priority']}\n\n")
            lines.append(f"**前置条件：** {case['precondition']}\n\n")
            if case['steps']:
                lines.append(f"**测试步骤：**\n\n{case['steps']}\n\n")
            lines.append(f"**预期结果：** {case['expected']}\n\n")
            lines.append("---\n\n")

        # 确定保存路径
        if not file_path:
            output_dir = os.path.join(os.getcwd(), 'outputs')
            os.makedirs(output_dir, exist_ok=True)
            filename = f"测试用例_{data.get('scenario', '场景')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            file_path = os.path.join(output_dir, filename)

        # 保存文件
        with open(file_path, 'w', encoding='utf-8') as f:
            f.writelines(lines)

        return file_path
