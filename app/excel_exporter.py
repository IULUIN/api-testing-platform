"""
Excel报告导出器
用于将压测结果导出为Excel格式
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime
import json


class ExcelReportExporter:
    """Excel报告导出器"""

    def __init__(self, result):
        """
        初始化导出器
        :param result: LoadTestResult对象
        """
        self.result = result
        self.wb = Workbook()

    def _create_summary_sheet(self):
        """创建概览页"""
        ws = self.wb.active
        ws.title = "压测概览"

        # 标题
        ws['A1'] = '压测报告'
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30

        # 基本信息
        row = 3
        info_data = [
            ('压测任务', self.result.load_test_name),
            ('目标接口', self.result.test_case_name),
            ('执行时间', self.result.executed_at.strftime('%Y-%m-%d %H:%M:%S')),
            ('', ''),
            ('并发用户数', f'{self.result.concurrent_users} 个'),
            ('持续时间', f'{self.result.duration} 秒'),
            ('Ramp-up时间', f'{self.result.ramp_up_time} 秒'),
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # 性能指标
        row += 1
        ws[f'A{row}'] = '性能指标'
        ws[f'A{row}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 25

        row += 1
        metrics_data = [
            ('总请求数', self.result.total_requests, '次'),
            ('成功请求', self.result.success_requests, '次'),
            ('失败请求', self.result.failed_requests, '次'),
            ('成功率', f'{(self.result.success_requests / self.result.total_requests * 100):.2f}' if self.result.total_requests > 0 else '0', '%'),
            ('', '', ''),
            ('TPS', f'{self.result.tps:.2f}', '次/秒'),
            ('平均响应时间', f'{self.result.avg_response_time * 1000:.2f}', 'ms'),
            ('最小响应时间', f'{self.result.min_response_time * 1000:.2f}', 'ms'),
            ('最大响应时间', f'{self.result.max_response_time * 1000:.2f}', 'ms'),
            ('', '', ''),
            ('P50 (中位数)', f'{self.result.p50_response_time * 1000:.2f}', 'ms'),
            ('P90', f'{self.result.p90_response_time * 1000:.2f}', 'ms'),
            ('P95', f'{self.result.p95_response_time * 1000:.2f}', 'ms'),
            ('P99', f'{self.result.p99_response_time * 1000:.2f}', 'ms'),
        ]

        for label, value, unit in metrics_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15

    def _create_timeseries_sheet(self):
        """创建时序数据页"""
        if not self.result.time_series_data:
            return

        ws = self.wb.create_sheet("时序数据")

        # 解析时序数据
        try:
            time_series = json.loads(self.result.time_series_data)
        except:
            return

        # 表头
        headers = ['时间(秒)', '活跃用户数', '每秒请求数', '累计请求数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 数据行
        for row_idx, data in enumerate(time_series, 2):
            ws.cell(row=row_idx, column=1, value=data.get('timestamp', 0))
            ws.cell(row=row_idx, column=2, value=data.get('active_users', 0))
            ws.cell(row=row_idx, column=3, value=data.get('requests', 0))
            ws.cell(row=row_idx, column=4, value=data.get('total_requests', 0))

        # 设置列宽
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 15

        # 添加图表
        if len(time_series) > 1:
            # 请求数趋势图
            chart = LineChart()
            chart.title = "每秒请求数趋势"
            chart.y_axis.title = "请求数"
            chart.x_axis.title = "时间(秒)"

            data = Reference(ws, min_col=3, min_row=1, max_row=len(time_series) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(time_series) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "F2")

    def export(self, filepath):
        """
        导出Excel文件
        :param filepath: 保存路径
        :return: 成功返回True，失败返回False
        """
        try:
            # 创建各个工作表
            self._create_summary_sheet()
            self._create_timeseries_sheet()

            # 保存文件
            self.wb.save(filepath)
            return True

        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False

